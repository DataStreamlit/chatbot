"""
or_aggregator.py
================
Merged replacement for **or_score.py** + **aggregate_or_data.py**.

Processes a folder of OR data collector workbooks (.xlsx / .xlsm) and writes
a **single** output Excel file with **two sheets**:

  Sheet 1 — "Specialty Level Data"   (previously Aggregated_Specialty_Level.xlsx)
  Sheet 2 — "Score"                  (previously OR_Score_Aggregated.xlsx)

Both sheets share the same embedded Key table (hospital → directorate mapping).
The Key table is also written as a third sheet "Key" for reference.

Usage
-----
  python or_aggregator.py --input "D:\\OR Files\\July 2025 Week 2" --output OR_Aggregated.xlsx
  python or_aggregator.py          # interactive mode — prompts for folder

Arguments
---------
  --input  / -i   Folder containing raw OR data collector files
  --output / -o   Output Excel file (default: OR_Aggregated.xlsx)
  --key    / -k   (Optional) Excel file with a Key sheet to override built-in mapping
"""

from __future__ import annotations

import argparse
import io
import logging
import os
import sys
import warnings
from datetime import datetime, timedelta

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS — SPECIALTY LEVEL DATA (Sheet 1)
# ══════════════════════════════════════════════════════════════════════════════

SPECIALTY_ROWS = 14

WAITING_TIME_ALIASES = [
    "OR Waiting Time",
    "OR Waiting time",
    "Waiting Time",
    "OR WaitingTime",
    "Waiting time",
]

SPEC_OUTPUT_COLS = [
    "Directorate",
    "Hospital Code",
    "Hospital Name",
    "Date",
    "Specialty",
    "Specialty available?\n(On Hold, Available, N/A)",
    "Total volume of patients on waiting list\n(any number or ND \"No Demand\")\nDon't write 0 or zero",
    "Volume of new patients added to the list\n(any number or ND \"No Demand\")\nDon't write 0 or zero",
    "Volume of patients with booked surgeries within the next 36 days\n(any number or ND \"No Demand\")\n",
    "Number of non scheduled cases due to shortage of supply\n(any number or ND \"No Demand\")\n",
    "Volume of patients without booked surgeries within the next 36 days\n(any number or ND \"No Demand\")\nDon't write 0 or zero",
    "OR session per week\n(per specialty)",
    "Average duration of the session (hours)\n ( 2, 4 , 6, 8) hours",
    "Total of Elective surgeries performed during this week",
    "Total Number of one day Surgeries performed during this week (According to MoH one day surgries list)",
    "Total of surgeries performed during this week",
    "Snapshot capture date \"Thursday\"\n(DD-MMM-YYYY)",
    "Date of 2nd next available slot appointment\n(DD-MMM-YYYY)",
    "Calculated days until 2nd Next available appointment",
    "Column16",
    "Number of Non-Emergency Functioning ORs",
    "Number of Non-Functioning ORs",
    "Number of Emergency Ors",
]

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS — SCORE (Sheet 2)
# ══════════════════════════════════════════════════════════════════════════════

SCORE_OUTPUT_COLS = [
    "Directorate",
    "Hospital Code",
    "Hospital Name",
    "Month",
    "Year",
    "Version",
    "Manual",
    "IT",
    "Score1",
    "Score2",
    "Score3",
    "Score4",
    "OR Utilization",
    "Elective surgery Volume Manual",
    "Emergency Surgery Volume Manual",
    "Or Utilization IT",
    "Surgical Cancellation IT",
    "Number of Non-Em Func OR WT",
    "Number of Non-Func OR WT",
    "Number of Em OR WT",
    "Elective surgery Volume IT",
    "Emergency Surgery Volume IT",
    "Elective Surgery Volume (Reconciled)",
    "Emergency Surgery Volume (Reconciled)",
]

KPI_MANUAL_ALIASES = [
    "OR KPI 1-4 & 6 Manual",
    "OR KPI 1-4 & 6 manual",
    "OR KPI 1-4 & 6 MANUAL",
    "KPI 1-4 & 6 Manual",
    "OR KPI 1-4 &6 Manual",
    "OR KPI 1-4&6 Manual",
]

KPI_IT_ALIASES = [
    "OR KPI 1-4 & 6 IT",
    "OR KPI 1-4 & 6 it",
    "OR KPI 1-4 & 6 IT ",
    "KPI 1-4 & 6 IT",
    "OR KPI 1-4 &6 IT",
]

# ══════════════════════════════════════════════════════════════════════════════
# EMBEDDED KEY TABLE  (hospital code → directorate)
# ══════════════════════════════════════════════════════════════════════════════

_EMBEDDED_KEY_TSV = """\
Hospital Code\tDirectorate
AA-GEN-1\tAl Ahsa Health Cluster
AA-GEN-2\tAl Ahsa Health Cluster
AA-GEN-3\tAl Ahsa Health Cluster
AA-GEN-4\tAl Ahsa Health Cluster
AA-GEN-5\tAl Ahsa Health Cluster
AA-GEN-6\tAl Ahsa Health Cluster
AA-GEN-7\tAl Ahsa Health Cluster
AA-GEN-8\tAl Ahsa Health Cluster
AA-MCH-1\tAl Ahsa Health Cluster
AS-GEN-1\tAsir RHD
AS-GEN-10\tAsir RHD
AS-GEN-11\tAsir RHD
AS-GEN-12\tAsir RHD
AS-GEN-13\tAsir RHD
AS-GEN-14\tAsir RHD
AS-GEN-15\tAsir RHD
AS-GEN-16\tAsir RHD
AS-GEN-17\tAsir RHD
AS-GEN-18\tAsir RHD
AS-GEN-2\tAsir RHD
AS-GEN-3\tAsir RHD
AS-GEN-4\tAsir RHD
AS-GEN-5\tAsir RHD
AS-GEN-6\tAsir RHD
AS-GEN-7\tAsir RHD
AS-GEN-8\tAsir RHD
AS-GEN-9\tAsir RHD
AS-MCH-1\tAsir RHD
AS-MCH-2\tAsir RHD
BA-GEN-1\tBaha RHD
BA-GEN-2\tBaha RHD
BA-GEN-3\tBaha RHD
BA-GEN-4\tBaha RHD
BA-GEN-5\tBaha RHD
BA-GEN-6\tBaha RHD
BA-GEN-7\tBaha RHD
BA-GEN-8\tBaha RHD
BH-GEN-1\tAsir RHD
BH-GEN-2\tAsir RHD
BH-GEN-3\tAsir RHD
BH-GEN-4\tAsir RHD
BH-GEN-5\tAsir RHD
BH-GEN-6\tAsir RHD
BH-MCH-1\tAsir RHD
C1-GEN-1\tRiyadh First Health Cluster
C1-GEN-10\tRiyadh First Health Cluster
C1-GEN-2\tRiyadh First Health Cluster
C1-GEN-3\tRiyadh First Health Cluster
C1-GEN-4\tRiyadh First Health Cluster
C1-GEN-5\tRiyadh First Health Cluster
C1-GEN-6\tRiyadh First Health Cluster
C1-GEN-7\tRiyadh First Health Cluster
C1-GEN-8\tRiyadh First Health Cluster
C1-GEN-9\tRiyadh First Health Cluster
C1-MDC-1\tRiyadh First Health Cluster
C2-GEN-1\tRiyadh Second Health Cluster
C2-GEN-2\tRiyadh Second Health Cluster
C2-GEN-3\tRiyadh Second Health Cluster
C2-GEN-4\tRiyadh Second Health Cluster
C2-GEN-5\tRiyadh Second Health Cluster
C2-MCH-1\tRiyadh Second Health Cluster
C2-MDC-1\tRiyadh Second Health Cluster
C3-GEN-1\tQassim Health Cluster
C3-GEN-10\tQassim Health Cluster
C3-GEN-11\tQassim Health Cluster
C3-GEN-12\tQassim Health Cluster
C3-GEN-13\tQassim Health Cluster
C3-GEN-14\tQassim Health Cluster
C3-GEN-15\tQassim Health Cluster
C3-GEN-16\tQassim Health Cluster
C3-GEN-17\tQassim Health Cluster
C3-GEN-2\tQassim Health Cluster
C3-GEN-3\tQassim Health Cluster
C3-GEN-4\tQassim Health Cluster
C3-GEN-5\tQassim Health Cluster
C3-GEN-6\tQassim Health Cluster
C3-GEN-7\tQassim Health Cluster
C3-GEN-8\tQassim Health Cluster
C3-GEN-9\tQassim Health Cluster
C3-MCH-1\tQassim Health Cluster
E1-GEN-1\tEastern Health Cluster
E1-GEN-10\tEastern Health Cluster
E1-GEN-11\tEastern Health Cluster
E1-GEN-12\tEastern Health Cluster
E1-GEN-13\tEastern Health Cluster
E1-GEN-14\tEastern Health Cluster
E1-GEN-15\tEastern Health Cluster
E1-GEN-16\tEastern Health Cluster
E1-GEN-17\tEastern Health Cluster
E1-GEN-18\tEastern Health Cluster
E1-GEN-2\tEastern Health Cluster
E1-GEN-3\tEastern Health Cluster
E1-GEN-4\tEastern Health Cluster
E1-GEN-5\tEastern Health Cluster
E1-GEN-6\tEastern Health Cluster
E1-GEN-7\tEastern Health Cluster
E1-GEN-8\tEastern Health Cluster
E1-GEN-9\tEastern Health Cluster
E1-MCH-1\tEastern Health Cluster
E1-MDC-1\tEastern Health Cluster
HA-GEN-1\tHail Health Cluster
HA-GEN-10\tHail Health Cluster
HA-GEN-11\tHail Health Cluster
HA-GEN-12\tHail Health Cluster
HA-GEN-2\tHail Health Cluster
HA-GEN-3\tHail Health Cluster
HA-GEN-4\tHail Health Cluster
HA-GEN-5\tHail Health Cluster
HA-GEN-6\tHail Health Cluster
HA-GEN-7\tHail Health Cluster
HA-GEN-8\tHail Health Cluster
HA-GEN-9\tHail Health Cluster
HA-MCH-1\tHail Health Cluster
HB-GEN-1\tHafer Al Batin RHD
HB-GEN-2\tHafer Al Batin RHD
HB-GEN-3\tHafer Al Batin RHD
HB-GEN-4\tHafer Al Batin RHD
HB-MCH-1\tHafer Al Batin RHD
JD-GEN-1\tJeddah Second Health Cluster
JD-GEN-2\tJeddah Second Health Cluster
JD-GEN-3\tJeddah Second Health Cluster
JD-GEN-4\tJeddah First Health Cluster
JD-GEN-5\tJeddah First Health Cluster
JD-GEN-6\tJeddah First Health Cluster
JD-GEN-7\tJeddah First Health Cluster
JD-GEN-8\tJeddah Second Health Cluster
JD-GEN-9\tJeddah First Health Cluster
JD-MCH-1\tJeddah Second Health Cluster
JF-GEN-1\tAl Jouf RHD
JF-GEN-2\tAl Jouf RHD
JF-GEN-3\tAl Jouf RHD
JF-GEN-4\tAl Jouf RHD
JF-GEN-5\tAl Jouf RHD
JF-GEN-6\tAl Jouf RHD
JF-GEN-7\tAl Jouf RHD
JF-GEN-8\tAl Jouf RHD
JF-GEN-9\tAl Jouf RHD
JF-MCH-1\tAl Jouf RHD
JZ-GEN-1\tJizan RHD
JZ-GEN-10\tJizan RHD
JZ-GEN-11\tJizan RHD
JZ-GEN-12\tJizan RHD
JZ-GEN-13\tJizan RHD
JZ-GEN-14\tJizan RHD
JZ-GEN-15\tJizan RHD
JZ-GEN-16\tJizan RHD
JZ-GEN-17\tJizan RHD
JZ-GEN-18\tJizan RHD
JZ-GEN-19\tJizan RHD
JZ-GEN-2\tJizan RHD
JZ-GEN-20\tJizan RHD
JZ-GEN-3\tJizan RHD
JZ-GEN-4\tJizan RHD
JZ-GEN-5\tJizan RHD
JZ-GEN-6\tJizan RHD
JZ-GEN-7\tJizan RHD
JZ-GEN-8\tJizan RHD
JZ-GEN-9\tJizan RHD
MD-GEN-1\tMadinah Health Cluster
MD-GEN-10\tMadinah Health Cluster
MD-GEN-11\tMadinah Health Cluster
MD-GEN-12\tMadinah Health Cluster
MD-GEN-13\tMadinah Health Cluster
MD-GEN-14\tMadinah Health Cluster
MD-GEN-15\tMadinah Health Cluster
MD-GEN-16\tMadinah Health Cluster
MD-GEN-2\tMadinah Health Cluster
MD-GEN-3\tMadinah Health Cluster
MD-GEN-4\tMadinah Health Cluster
MD-GEN-6\tMadinah Health Cluster
MD-GEN-7\tMadinah Health Cluster
MD-GEN-8\tMadinah Health Cluster
MD-GEN-9\tMadinah Health Cluster
MD-MCH-1\tMadinah Health Cluster
MD-MDC-1\tMadinah Health Cluster
NB-GEN-1\tNorthern Border RHD
NB-GEN-2\tNorthern Border RHD
NB-GEN-3\tNorthern Border RHD
NB-GEN-4\tNorthern Border RHD
NB-GEN-5\tNorthern Border RHD
NB-GEN-6\tNorthern Border RHD
NB-GEN-7\tNorthern Border RHD
NB-MCH-1\tNorthern Border RHD
NB-MCH-2\tNorthern Border RHD
NJ-GEN-1\tNajran RHD
NJ-GEN-2\tNajran RHD
NJ-GEN-3\tNajran RHD
NJ-GEN-4\tNajran RHD
NJ-GEN-5\tNajran RHD
NJ-GEN-6\tNajran RHD
NJ-GEN-7\tNajran RHD
NJ-GEN-8\tNajran RHD
NJ-MCH-1\tNajran RHD
NJ-MEN-1\tNajran RHD
QN-GEN-1\tMakkah Health Cluster
QN-GEN-2\tMakkah Health Cluster
QN-GEN-3\tMakkah Health Cluster
QN-GEN-4\tMakkah Health Cluster
QN-GEN-5\tMakkah Health Cluster
QR-GEN-1\tAl Jouf RHD
QR-GEN-2\tAl Jouf RHD
QR-GEN-3\tAl Jouf RHD
RH-GEN-1\tRiyadh First Health Cluster
RH-GEN-10\tRiyadh Third Health Cluster
RH-GEN-11\tRiyadh Third Health Cluster
RH-GEN-12\tRiyadh Third Health Cluster
RH-GEN-13\tRiyadh Third Health Cluster
RH-GEN-14\tRiyadh Second Health Cluster
RH-GEN-15\tRiyadh Third Health Cluster
RH-GEN-16\tRiyadh Third Health Cluster
RH-GEN-17\tRiyadh Third Health Cluster
RH-GEN-18\tRiyadh Third Health Cluster
RH-GEN-19\tRiyadh Third Health Cluster
RH-GEN-2\tRiyadh Second Health Cluster
RH-GEN-20\tRiyadh Third Health Cluster
RH-GEN-21\tRiyadh Third Health Cluster
RH-GEN-22\tRiyadh Third Health Cluster
RH-GEN-23\tRiyadh Third Health Cluster
RH-GEN-3\tRiyadh Second Health Cluster
RH-GEN-5\tRiyadh First Health Cluster
RH-GEN-6\tRiyadh First Health Cluster
RH-GEN-7\tRiyadh First Health Cluster
RH-GEN-9\tRiyadh First Health Cluster
RH-MCH-1\tRiyadh First Health Cluster
TB-GEN-1\tTabuk RHD
TB-GEN-10\tTabuk RHD
TB-GEN-11\tTabuk RHD
TB-GEN-2\tTabuk RHD
TB-GEN-3\tTabuk RHD
TB-GEN-4\tTabuk RHD
TB-GEN-5\tTabuk RHD
TB-GEN-6\tTabuk RHD
TB-GEN-7\tTabuk RHD
TB-GEN-8\tTabuk RHD
TB-GEN-9\tTabuk RHD
TB-MCH-1\tTabuk RHD
TF-GEN-1\tTaif RHD
TF-GEN-10\tTaif RHD
TF-GEN-11\tTaif RHD
TF-GEN-12\tTaif RHD
TF-GEN-13\tTaif RHD
TF-GEN-2\tTaif RHD
TF-GEN-3\tTaif RHD
TF-GEN-4\tTaif RHD
TF-GEN-5\tTaif RHD
TF-GEN-6\tTaif RHD
TF-GEN-7\tTaif RHD
TF-GEN-8\tTaif RHD
TF-GEN-9\tTaif RHD
TF-MCH-1\tTaif RHD
W1-GEN-1\tMakkah Health Cluster
W1-GEN-2\tMakkah Health Cluster
W1-GEN-3\tMakkah Health Cluster
W1-GEN-4\tMakkah Health Cluster
W1-GEN-5\tMakkah Health Cluster
W1-GEN-6\tMakkah Health Cluster
W1-GEN-7\tMakkah Health Cluster
W1-MCH-1\tMakkah Health Cluster
W1-MDC-1\tMakkah Health Cluster
"""

# ══════════════════════════════════════════════════════════════════════════════
# EMBEDDED HOSPITAL NAME TABLE  (hospital code → hospital name)
# ══════════════════════════════════════════════════════════════════════════════

_HOSPITAL_NAME_TSV = """\
Hospital Code\tHospital Name
AA-GEN-1\tKing Fahad Central Hospital In Hafouf
AA-GEN-2\tPrince Saud Bin Jalloway Hospital
AA-GEN-3\tAl-Jabr ENT and Eye Hospital
AA-MCH-1\tMaternity & Children Hospital in Hassa
AA-GEN-4\tPrince Sultan Center
AA-GEN-5\tKing Faisal Hospital in Hassa
AA-GEN-6\tAlamaraan Hospital in Hassa
AA-GEN-8\tMadinat Alayon General Hospital
AS-GEN-1\tAsir Central Hospital
AS-GEN-10\tUhod Rafidah General Hospital
AS-GEN-2\tMahayel General Hospital
AS-GEN-3\tKhamis Mushayt General Hospital
AS-GEN-4\tSarat Ubaida General Hospital
AS-GEN-5\tDhahran Al-JaNOab General Hospital
AS-GEN-6\tAl-Namas General Hospital
AS-GEN-7\tBllsamar General Hospital
AS-GEN-8\tAl-Majardah General Hospital
AS-GEN-9\tRejal Almaa General Hospital
AS-MCH-1\tKhamis Mushayt Maternity Hospital
AS-MCH-2\tMaternity & Children's Hospital in Abha
AS-GEN-14\tBallahmar Hospital
BA-GEN-1\tPrince Meshari Bin Saud General Hospital
BA-GEN-2\tKing Fahad Hospital in Al-Bahah
BA-GEN-3\tAlmakhawa General Hospital
BA-GEN-4\tQlwa General Hospital
BA-GEN-5\tAlmndiq General Hospital
BA-GEN-7\tAlaqiq General Hospital
BA-GEN-8\tAlqraa Hospital
BH-GEN-1\tKing Abdullah Hospital in Beshah
BH-GEN-2\tTathleeth General Hospital
BH-GEN-4\tSabt Alalaya General Hospital
BH-GEN-5\tBasher General Hospital
BH-MCH-1\tMaternity & Children's Hospital in Bishah
HA-GEN-1\tKing Khaled Hospital in Hail
HA-GEN-2\tHail General Hospital
HA-GEN-3\tKing Salman Specialized Hospital
HA-MCH-1\tMaternity and Children Hospital
HA-GEN-10\tAlshnan General Hospital
HA-GEN-11\tAlhaeit Hospital
HA-GEN-4\tAlsleimi General Hospital
HA-GEN-5\tAlshamly General Hospital
HA-GEN-6\tMuqaq General Hospital
HA-GEN-7\tAlbaqa'a General Hospital
HA-GEN-8\tSameraa General Hospital
HA-GEN-9\tAlghazalah Hospital
HA-GEN-12\tCardiac center in hail
HB-GEN-1\tKing Khaled Hospital in Hafer Al-Baten
HB-GEN-2\tHafer Al Batin Central hospital
HB-MCH-1\tMaternity and Children Hospital
JD-GEN-1\tKing Fahad Hospital
JD-GEN-2\tKing Abdullah Medical Complex (NOrthern Jeddah)
JD-GEN-3\tRabigh General Hospital
JD-GEN-4\tEast Jeddah Hospital
JD-GEN-5\tAl-Thagher General Hospital
JD-GEN-6\tKing Abdul Aziz Hospital and Oncology Center in Jeddah
JD-GEN-7\tAdhom General Hospital
JD-GEN-8\tEye Hospital in Jeddah
JD-GEN-9\tAl-leith Hospital
JD-MCH-1\tMaternity & Children's Hospital in the north
JF-GEN-1\tPrince Moteb bin Abdulaziz Hospital
JF-GEN-2\tTabarjal General Hospital
JF-GEN-3\tDomat Al-Jandal General Hospital
JF-GEN-4\tKing Abdulaziz Specialist Hospital in Jouf
JF-MCH-1\tMaternity & Children Hospital in Skaka Jouf
JF-GEN-5\tSoder General Hospital
JF-GEN-8\tCardiology Hospital
JZ-GEN-1\tPrince Mohammed bin Nasser Hospital
JZ-GEN-2\tSametah General Hospital
JZ-GEN-3\tSabya General Hospital
JZ-GEN-4\tAbu Arish General Hospital
JZ-GEN-5\tBesh General hospital
JZ-GEN-6\tKing Fahd Central Hospital in Jazan
JZ-GEN-10\tAlaarda General Hospital
JZ-GEN-12\tAldarb General Hospital
JZ-GEN-13\tAltawal General Hospital
JZ-GEN-14\tAhad Almassarha General Hospital
JZ-GEN-15\tAlmossem General Hospital
JZ-GEN-16\tBani Malik Hospital
JZ-GEN-17\tDamad General Hospital
JZ-GEN-19\tAlkhobah Hospital (Alharth)
JZ-GEN-20\tAledabi Hospital
JZ-GEN-7\tJazan General Hospital
JZ-GEN-8\tAlfursan General Hospital
JZ-GEN-9\tFefa General Hospital
MD-GEN-1\tKing Fahad Hospital
MD-GEN-2\tYanbu General Hospital
MD-GEN-3\tUhod General Hospital
MD-GEN-4\tPrince Abdul Mohsen Hospital Al-Ola
MD-GEN-6\tKhayber General Hospital
MD-MDC-1\tKing Salman Bin Abdulaziz Medical City
MD-GEN-10\tBader General Hospital
MD-GEN-11\tAlhnakiyah General Hospital
MD-GEN-12\tAlaeis General Hospital
MD-GEN-13\tWadi Alfarie General Hospital
MD-GEN-14\tAlhmnah Hospital
MD-GEN-15\tAlhasso Hospital
MD-GEN-16\tYanbu Alnakhel Hospital
MD-GEN-7\tCardiology Hospital
MD-GEN-8\tAlmiqat General Hospital
MD-GEN-9\tAlmahd General Hospital
NB-GEN-1\tArar Central Hospital
NB-GEN-2\tTarif General Hospital
NB-GEN-3\tPrince Abdulaziz Bin Mosaad Bin Jalloway Hospital
NB-GEN-4\tRafha General Hospital
NB-MCH-1\tMaternity & Children Hospital in Arar
NB-GEN-5\tAlaoiqilh Hospital
NB-GEN-6\tSho'bah General Hospital
NB-GEN-7\tJadidah Arar Hospital
NB-MCH-2\tMaternity & Children Hospital in Rafha
NJ-GEN-1\tKing Khaled Hospital in Najran
NJ-GEN-2\tNajran General Hospital
NJ-GEN-3\tSharorah General Hospital
NJ-MCH-1\tMaternity and Children Hospital
QN-GEN-1\tAl-Qunfudah General Hospital
QN-GEN-2\tSouth Qunfudah Hospital
QN-GEN-3\tTripan Hospital
QN-GEN-4\tAL-Mezailef General Hospital
QN-GEN-5\tNamirah General Hospital
QR-GEN-1\tAl-Qrayat General Hospital
RH-GEN-1\tKing Khalid Hospital Kharj
RH-GEN-10\tDawadmi Hospital
RH-GEN-11\tShaqraa Hospital
RH-GEN-12\tAfif Hospital
RH-GEN-13\tHemaymlah Hospital
RH-GEN-14\tHotat Sedir Hospital
RH-GEN-2\tKing Khalid Hospital Magmah
RH-GEN-3\tZulfi Hospital
RH-GEN-5\tGoeyah Hospital
RH-GEN-6\tWadi Adwaser Hospital
RH-GEN-7\tAflaj Hospital
RH-GEN-9\tHotat Tamem Hospital
RH-MCH-1\tChildren and Delivery Hospital Kharj
RH-GEN-15\tAlderiyadh Hospital
RH-GEN-16\tAlthadek Hospital
RH-GEN-17\tAlsajer Hospital
RH-GEN-18\tAlnafee Hospital
RH-GEN-19\tAlwathlyan Hospital
RH-GEN-20\tAlrafyie Hospital in Jamash
RH-GEN-22\tAldrma Hospital
TB-GEN-1\tKing Khalid Hospital in Tabuk
TB-GEN-2\tAlwagh General Hospital
TB-GEN-3\tTemaa General Hospital
TB-GEN-4\tHagh General Hospital
TB-GEN-5\tOmlog General Hospital
TB-GEN-6\tDubaa General Hospital
TB-GEN-7\tKing Fahad Specialized Hospital in Tabuk
TB-MCH-1\tChildren & Delivery Hospital in Tabuk
TF-GEN-1\tKing Faisal Hospital in Taif
TF-GEN-2\tKing Abdulaziz Specialist Hospital in Taif
TF-MCH-1\tMaternity Hospital in Taif
TF-GEN-10\tThelim Hospital
TF-GEN-11\tQiya Balhareth Hospital
TF-GEN-13\tAlmahani Hospital
TF-GEN-3\tAlkhurma General Hospital
TF-GEN-4\tMeisan Balhareth General Hospital
TF-GEN-5\tRnyah General Hospital
TF-GEN-6\tTurba General Hospital
TF-GEN-7\tAlkhurei Bni Malik General Hospital
TF-GEN-8\tSlsahin Bin Saad General Hospital
TF-GEN-9\tAlmoya Hospital
TF-GEN-12\tUm Aldoum Hospital
C1-GEN-1\tKing Salman Bin Abdulaziz Hospital
C1-GEN-2\tAl Iman General Hospital
C1-GEN-3\tAl Imam Abdul Rahman Al Faisal Hospital
C1-MDC-1\tGeneral Hospital King Saud Medical City
C1-GEN-10\tAlrein Hospital
C1-GEN-4\tAlsuliel Hospital
C1-GEN-5\tRweidha Alaard Hospital
C1-GEN-6\tAlmuzahmiyah Hospital
C1-GEN-7\tPrince Salman bin Mohammed Hospital (Aldlim)
C1-GEN-8\tAlhuraig Hospital
C1-GEN-9\tAlkhaseira Hospital
C2-GEN-1\tPrince Mohammed bin Abdul Aziz Hospital
C2-MCH-1\tAl Yamamah Hospital
C2-MDC-1\tKing Fahad Medical City
C2-GEN-2\tAlrumah Hospital
C2-GEN-3\tTumeir Hospital
C2-GEN-4\tAlghat Hospital
C2-GEN-5\tAlartauya Hospital
C3-GEN-1\tBuraydah Central Hospital
C3-GEN-2\tKing Saud Hospital in Oniazah
C3-GEN-3\tAl-Rass General Hospital
C3-GEN-4\tAl-Bikeriah General Hospital
C3-GEN-5\tAl-Mothneb General Hospital
C3-GEN-6\tAl-Badaa General Hospital
C3-GEN-7\tKing Fahd Specialist Hospital in Buraydah
C3-MCH-1\tMaternity and Children Hospital in Qassim
C3-GEN-11\tAyoon Aljwaa General Hospital
C3-GEN-12\tRyad alkhubraa General Hospital
C3-GEN-13\tAkla alsagoor General Hospital
C3-GEN-14\tdarba General Hospital
C3-GEN-15\tAlasyah Hospital
C3-GEN-16\tAlnabhaniya Hospital
C3-GEN-17\tAlqwara Hospital
C3-GEN-8\tCardiology Hospital
E1-GEN-1\tDammam Medical Complex
E1-GEN-2\tAl Qatif Central Hospital
E1-GEN-3\tAl-Jubail General Hospital
E1-GEN-4\tAl Khafji General Hospital
E1-MCH-1\tMaternity and Children Hospital in Dammam
E1-MDC-1\tKing Fahad Specialist Hospital in Dammam (KFSH-D)
E1-GEN-10\tAlrafyie General Hospital
E1-GEN-13\tPrince Sultan Hospial in Bmleja
E1-GEN-14\tPrince Sultan Hospital in Bariera
E1-GEN-15\tAlqarya alalya Hospital
E1-GEN-16\tRas Tanoura Hospital
E1-GEN-18\tSalwa Hospital
E1-GEN-5\tBabtin Hospital
E1-GEN-6\tPrince Mohammed bin Fahad General Hospital
E1-GEN-7\tAlthahran Altkhasusi Hospital
E1-GEN-8\tAlnaariyah General Hospital
E1-GEN-9\tAlbakek General Hospital
W1-GEN-1\tKing Abdulaziz Hospital in Makkah
W1-GEN-2\tKing Faisal Hospital in Makkah
W1-GEN-3\tHera General Hospital
W1-GEN-4\tAl-NOor Specialist Hospital
W1-MCH-1\tMaternity & Children's Hospital in Makkah
W1-MDC-1\tKing Abdullah Medical City
W1-GEN-5\tAlkhulais Hospital
W1-GEN-6\tAlkamel Hospital
"""

# ══════════════════════════════════════════════════════════════════════════════
# SHARED KEY HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def get_embedded_key() -> tuple[dict, pd.DataFrame]:
    """Parse the embedded TSV and return (key_map, key_df)."""
    df = pd.read_csv(io.StringIO(_EMBEDDED_KEY_TSV), sep="\t", dtype=str)
    df.columns = ["Hospital Code", "Directorate"]
    df = df.dropna(subset=["Hospital Code", "Directorate"])
    df["Hospital Code"] = df["Hospital Code"].str.strip().str.upper()
    df["Directorate"]   = df["Directorate"].str.strip()
    key_map = dict(zip(df["Hospital Code"], df["Directorate"]))
    return key_map, df


def get_hospital_name_map() -> dict:
    """Return the embedded hospital code → hospital name lookup dict."""
    df = pd.read_csv(io.StringIO(_HOSPITAL_NAME_TSV), sep="\t", dtype=str)
    df["Hospital Code"] = df["Hospital Code"].str.strip().str.upper()
    df["Hospital Name"] = df["Hospital Name"].str.strip()
    return dict(zip(df["Hospital Code"], df["Hospital Name"]))


def load_key_from_file(
    key_path: str | None = None,
    search_folder: str | None = None,
) -> tuple[dict | None, pd.DataFrame | None]:
    """
    Try to load an external Key sheet to override the embedded key.
    Priority: explicit --key path → any file in search_folder with a 'Key' sheet.
    """
    candidates = []
    if key_path:
        candidates.append(key_path)
    if search_folder and os.path.isdir(search_folder):
        for fname in sorted(os.listdir(search_folder)):
            if fname.endswith((".xlsx", ".xlsm")) and not fname.startswith("~$"):
                candidates.append(os.path.join(search_folder, fname))

    for path in candidates:
        if not os.path.exists(path):
            continue
        try:
            xl = pd.ExcelFile(path)
            if "Key" not in xl.sheet_names:
                continue
            key_df = pd.read_excel(path, sheet_name="Key", header=0)
            key_df = key_df.iloc[:, :2].copy()
            key_df.columns = ["Hospital Code", "Directorate"]
            key_df = key_df.dropna(subset=["Hospital Code", "Directorate"])
            key_df["Hospital Code"] = key_df["Hospital Code"].astype(str).str.strip().str.upper()
            key_df = key_df[key_df["Hospital Code"].str.contains("-")]
            if len(key_df) == 0:
                continue
            key_map = dict(zip(key_df["Hospital Code"], key_df["Directorate"]))
            logging.info(
                f"Key sheet override from: {os.path.basename(path)} ({len(key_map)} hospitals)"
            )
            return key_map, key_df
        except Exception:
            continue
    return None, None


# ══════════════════════════════════════════════════════════════════════════════
# SHEET NAME RESOLUTION HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _resolve_sheet(sheet_names: list[str], aliases: list[str], fuzzy_keyword: str) -> str | None:
    for alias in aliases:
        if alias in sheet_names:
            return alias
    for name in sheet_names:
        if fuzzy_keyword.lower() in name.lower():
            return name
    return None


def find_waiting_time_sheet(sheet_names: list[str]) -> str | None:
    for alias in WAITING_TIME_ALIASES:
        if alias in sheet_names:
            return alias
    for name in sheet_names:
        if "waiting" in name.lower():
            return name
    return None


def find_kpi_manual_sheet(sheet_names: list[str]) -> str | None:
    return _resolve_sheet(sheet_names, KPI_MANUAL_ALIASES, "manual")


def find_kpi_it_sheet(sheet_names: list[str]) -> str | None:
    return _resolve_sheet(sheet_names, KPI_IT_ALIASES, "kpi")


# ══════════════════════════════════════════════════════════════════════════════
# CELL READING HELPERS  (used by Score processing)
# ══════════════════════════════════════════════════════════════════════════════

def _cell(ws, row: int, col: int):
    """Read a single openpyxl cell value (1-based). Returns None on error."""
    try:
        val = ws.cell(row=row, column=col).value
        return val
    except Exception:
        return None


def _safe_num(val):
    """Convert cell value to number or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s.startswith("#") or s == "":
        return None
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, OverflowError):
        return None


def _safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ══════════════════════════════════════════════════════════════════════════════
# DATE HELPERS  (used by Specialty Level processing)
# ══════════════════════════════════════════════════════════════════════════════

def nth_thursday_of_month(month_str: str, year_str: str, week_num: int) -> datetime:
    dt = datetime.strptime(f"1 {month_str} {year_str}", "%d %B %Y")
    days_ahead = 3 - dt.weekday()
    if days_ahead <= 0:
        days_ahead += 7
    first_thursday = dt + timedelta(days=days_ahead)
    return first_thursday + timedelta(weeks=week_num)


def fallback_date_from_summary(wb_path: str) -> datetime | None:
    try:
        ss = pd.read_excel(wb_path, sheet_name="Summary Sheet", header=None)
        month    = str(ss.iloc[7, 2]).strip()
        year_raw = ss.iloc[8, 2]
        week_raw = ss.iloc[23, 2]
        year  = str(int(float(year_raw)))
        week  = int(float(week_raw))
        return nth_thursday_of_month(month, year, week)
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# FILE DETECTION
# ══════════════════════════════════════════════════════════════════════════════

def is_or_data_file(filepath: str) -> tuple[bool, str | None]:
    """Return (True, wt_sheet_name) if file has Summary Sheet + Waiting Time sheet."""
    try:
        sheets = pd.ExcelFile(filepath).sheet_names
        if "Summary Sheet" not in sheets:
            return False, None
        wt_sheet = find_waiting_time_sheet(sheets)
        return (True, wt_sheet) if wt_sheet else (False, None)
    except Exception:
        return False, None


def is_or_score_file(filepath: str) -> tuple[bool, dict]:
    """Return (True, sheet_map) if file has Summary Sheet + at least one KPI sheet."""
    try:
        sheets = pd.ExcelFile(filepath).sheet_names
        if "Summary Sheet" not in sheets:
            return False, {}
        kpi_manual = find_kpi_manual_sheet(sheets)
        if kpi_manual is None:
            return False, {}
        kpi_it       = find_kpi_it_sheet(sheets)
        waiting_time = find_waiting_time_sheet(sheets)
        return True, {
            "kpi_manual":   kpi_manual,
            "kpi_it":       kpi_it,
            "waiting_time": waiting_time,
        }
    except Exception:
        return False, {}


# ══════════════════════════════════════════════════════════════════════════════
# SPECIALTY LEVEL DATA — FILE PROCESSING
# ══════════════════════════════════════════════════════════════════════════════

def read_hospital_code(filepath: str) -> str:
    df = pd.read_excel(filepath, sheet_name="Summary Sheet", header=None)
    return str(df.iloc[9, 2]).strip().upper()


def read_waiting_time(filepath: str, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=0)
    return df.iloc[:SPECIALTY_ROWS].copy()


def process_spec_file(
    filepath: str, key_map: dict, unknown_codes: set, name_map: dict = None
) -> list[dict]:
    """Process one OR data collector file → list of specialty-level row dicts."""
    rows = []
    fname = os.path.basename(filepath)

    try:
        ok, wt_sheet = is_or_data_file(filepath)
        if not ok:
            logging.warning(f"SKIPPED (no required sheets): {fname}")
            return []

        hospital_code = read_hospital_code(filepath)
        wt_df         = read_waiting_time(filepath, wt_sheet)
        directorate   = key_map.get(hospital_code, "")
        hospital_name = (name_map or {}).get(hospital_code, "")

        if not directorate:
            unknown_codes.add(hospital_code)
            logging.warning(
                f"Unknown hospital code '{hospital_code}' in {fname} — Directorate blank"
            )

        # Snapshot Thursday: try sheet col 12 first, fall back to Summary Sheet
        snapshot_thursday = None
        for val in wt_df.iloc[:, 12]:
            if pd.notna(val):
                snapshot_thursday = val
                break

        if snapshot_thursday is None:
            snapshot_thursday = fallback_date_from_summary(filepath)
            if snapshot_thursday:
                logging.info(
                    f"Date fallback used for {hospital_code}: "
                    f"{snapshot_thursday.date() if hasattr(snapshot_thursday, 'date') else snapshot_thursday}"
                )
            else:
                logging.warning(f"Could not derive date for {hospital_code} in {fname}")

        # Date column = Thursday + 3 days (Sunday), filled for ALL rows
        if snapshot_thursday is not None:
            try:
                week_sunday = pd.Timestamp(snapshot_thursday) + timedelta(days=3)
            except Exception:
                week_sunday = None
        else:
            week_sunday = None

        for i in range(SPECIALTY_ROWS):
            row       = wt_df.iloc[i]
            specialty = row.iloc[0]

            if pd.isna(specialty) or str(specialty).strip() in ("", "x"):
                continue

            specialty_available = row.iloc[1]
            row_has_data = pd.notna(specialty_available)

            row_thursday = row.iloc[12] if pd.notna(row.iloc[12]) else (
                snapshot_thursday if row_has_data else None
            )

            if row_has_data and row_thursday is not None:
                try:
                    row_sunday = pd.Timestamp(row_thursday) + timedelta(days=3)
                except Exception:
                    row_sunday = week_sunday
            else:
                row_sunday = week_sunday

            raw_calc = row.iloc[14]
            try:
                calc_days = int(round(float(raw_calc))) if pd.notna(raw_calc) else None
            except Exception:
                calc_days = None

            # Keep null as null for "without booked 36d"
            raw_wob = row.iloc[6]
            wob_val = raw_wob if pd.notna(raw_wob) else None

            rows.append({
                "Directorate":       directorate,
                "Hospital Code":     hospital_code,
                "Hospital Name":     hospital_name,
                "Date":              row_sunday,
                "Specialty":         specialty,
                SPEC_OUTPUT_COLS[5]:  specialty_available,
                SPEC_OUTPUT_COLS[6]:  row.iloc[2],
                SPEC_OUTPUT_COLS[7]:  row.iloc[3],
                SPEC_OUTPUT_COLS[8]:  row.iloc[4],
                SPEC_OUTPUT_COLS[9]:  row.iloc[5],
                SPEC_OUTPUT_COLS[10]: wob_val,
                SPEC_OUTPUT_COLS[11]: row.iloc[7],
                SPEC_OUTPUT_COLS[12]: row.iloc[8],
                SPEC_OUTPUT_COLS[13]: row.iloc[9],
                SPEC_OUTPUT_COLS[14]: row.iloc[10],
                SPEC_OUTPUT_COLS[15]: row.iloc[11],
                SPEC_OUTPUT_COLS[16]: row_thursday,
                SPEC_OUTPUT_COLS[17]: row.iloc[13],
                SPEC_OUTPUT_COLS[18]: calc_days,
                SPEC_OUTPUT_COLS[19]: row.iloc[15] if len(row) > 15 else None,
                SPEC_OUTPUT_COLS[20]: row.iloc[16] if len(row) > 16 else None,
                SPEC_OUTPUT_COLS[21]: row.iloc[17] if len(row) > 17 else None,
                SPEC_OUTPUT_COLS[22]: row.iloc[18] if len(row) > 18 else None,
            })

        label = directorate if directorate else "NO DIRECTORATE"
        msg = f"  OK  {fname}: {hospital_code} ({label}) — {len(rows)} specialty rows"
        if wt_sheet != "OR Waiting Time":
            msg += f" [sheet alias: '{wt_sheet}']"
        logging.info(msg)
        print(msg)

    except Exception as e:
        msg = f"  ERR {fname}: {e}"
        logging.error(msg)
        print(msg)

    return rows


# ══════════════════════════════════════════════════════════════════════════════
# SCORE — FILE PROCESSING
# ══════════════════════════════════════════════════════════════════════════════

def process_score_file(filepath: str, key_map: dict, unknown_codes: set) -> dict | None:
    """Process one OR data collector file → single score row dict or None."""
    fname = os.path.basename(filepath)

    try:
        ok, sheet_map = is_or_score_file(filepath)
        if not ok:
            logging.warning(f"SKIPPED (no KPI sheets): {fname}")
            return None

        kpi_manual_name   = sheet_map["kpi_manual"]
        kpi_it_name       = sheet_map["kpi_it"]
        waiting_time_name = sheet_map["waiting_time"]

        wb = openpyxl.load_workbook(
            filepath, read_only=True, data_only=True, keep_vba=False
        )

        # ── Summary Sheet ──────────────────────────────────────────────────────
        ss = wb["Summary Sheet"]
        hosp_code  = _safe_str(_cell(ss, 10, 3))
        hosp_name  = _safe_str(_cell(ss, 11, 3))
        month      = _safe_str(_cell(ss, 8,  3))
        year       = _safe_num(_cell(ss, 9,  3))
        version    = _safe_str(_cell(ss, 1,  1))
        manual_flg = _safe_str(_cell(ss, 3,  2))
        it_flg     = _safe_str(_cell(ss, 4,  2))
        score1     = _cell(ss, 5, 1)
        score2     = _cell(ss, 6, 1)
        score3     = _cell(ss, 7, 1)
        score4     = _cell(ss, 8, 1)

        if hosp_code:
            hosp_code = hosp_code.upper()
        else:
            hosp_code = "UNKNOWN"

        directorate = key_map.get(hosp_code, "")
        if not directorate:
            unknown_codes.add(hosp_code)
            logging.warning(
                f"Unknown hospital code '{hosp_code}' in {fname} — Directorate blank"
            )

        # ── OR KPI 1-4 & 6 Manual ─────────────────────────────────────────────
        utilization_m   = None
        elective_vol_m  = None
        emergency_vol_m = None

        if kpi_manual_name and kpi_manual_name in wb.sheetnames:
            km = wb[kpi_manual_name]
            utilization_m   = _safe_num(_cell(km, 2, 60))   # BH2
            elective_vol_m  = _safe_num(_cell(km, 2, 64))   # BL2
            emergency_vol_m = _safe_num(_cell(km, 2, 66))   # BN2
        else:
            logging.warning(f"KPI Manual sheet not found in {fname}")

        # ── OR KPI 1-4 & 6 IT ─────────────────────────────────────────────────
        utilization_it   = None
        cancellation_it  = None
        elective_vol_it  = None
        emergency_vol_it = None

        if kpi_it_name and kpi_it_name in wb.sheetnames:
            ki = wb[kpi_it_name]
            utilization_it   = _safe_num(_cell(ki, 3, 27))  # AA3
            cancellation_it  = _safe_num(_cell(ki, 3, 28))  # AB3
            elective_vol_it  = _safe_num(_cell(ki, 3, 21))  # U3
            emergency_vol_it = _safe_num(_cell(ki, 3, 22))  # V3
        else:
            logging.warning(f"KPI IT sheet not found in {fname}")

        # ── OR Waiting Time ────────────────────────────────────────────────────
        func_or  = None
        nfunc_or = None
        em_or    = None

        if waiting_time_name and waiting_time_name in wb.sheetnames:
            wt = wb[waiting_time_name]
            func_or  = _safe_num(_cell(wt, 3, 17))   # Q3
            nfunc_or = _safe_num(_cell(wt, 3, 18))   # R3
            em_or    = _safe_num(_cell(wt, 3, 19))   # S3
        else:
            logging.warning(f"OR Waiting Time sheet not found in {fname}")

        wb.close()

        row = {
            "Directorate":                     directorate,
            "Hospital Code":                   hosp_code,
            "Hospital Name":                   hosp_name,
            "Month":                           month,
            "Year":                            year,
            "Version":                         version,
            "Manual":                          manual_flg,
            "IT":                              it_flg,
            "Score1":                          score1,
            "Score2":                          score2,
            "Score3":                          score3,
            "Score4":                          score4,
            "OR Utilization":                  utilization_m,
            "Elective surgery Volume Manual":  elective_vol_m,
            "Emergency Surgery Volume Manual": emergency_vol_m,
            "Or Utilization IT":               utilization_it,
            "Surgical Cancellation IT":        cancellation_it,
            "Number of Non-Em Func OR WT":     func_or,
            "Number of Non-Func OR WT":        nfunc_or,
            "Number of Em OR WT":              em_or,
            "Elective surgery Volume IT":      elective_vol_it,
            "Emergency Surgery Volume IT":     emergency_vol_it,
            # Reconciled: take IT if IT > Manual (or Manual is None)
            "Elective Surgery Volume (Reconciled)": (
                elective_vol_it
                if (elective_vol_m is not None and elective_vol_it is not None
                    and elective_vol_m < elective_vol_it)
                else (elective_vol_it if elective_vol_m is None else elective_vol_m)
            ),
            "Emergency Surgery Volume (Reconciled)": (
                emergency_vol_it
                if (emergency_vol_m is not None and emergency_vol_it is not None
                    and emergency_vol_m < emergency_vol_it)
                else (emergency_vol_it if emergency_vol_m is None else emergency_vol_m)
            ),
        }

        label = directorate if directorate else "NO DIRECTORATE"
        msg = f"  OK  {fname}: {hosp_code} ({label}) — score row"
        logging.info(msg)
        print(msg)
        return row

    except Exception as e:
        msg = f"  ERR {fname}: {e}"
        logging.error(msg)
        print(msg)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# FOLDER AGGREGATION
# ══════════════════════════════════════════════════════════════════════════════

def aggregate_folder(
    folder_path: str,
    key_map: dict,
    name_map: dict = None,
) -> tuple[pd.DataFrame, pd.DataFrame, set]:
    """
    Scan folder, process all valid OR data collector files.

    Returns
    -------
    spec_df      : DataFrame with SPEC_OUTPUT_COLS  (Specialty Level Data)
    score_df     : DataFrame with SCORE_OUTPUT_COLS (Score)
    unknown_codes: set of hospital codes not found in key_map
    """
    spec_rows     = []
    score_rows    = []
    unknown_codes = set()

    files = sorted([
        f for f in os.listdir(folder_path)
        if f.endswith((".xlsx", ".xlsm")) and not f.startswith("~$")
    ])

    msg = f"\nFound {len(files)} Excel file(s) in folder. Scanning...\n"
    logging.info(msg)
    print(msg)

    for fname in files:
        fpath = os.path.join(folder_path, fname)

        # ── Specialty Level Data ───────────────────────────────────────────────
        ok_spec, _ = is_or_data_file(fpath)
        if ok_spec:
            rows = process_spec_file(fpath, key_map, unknown_codes, name_map=name_map)
            spec_rows.extend(rows)
        else:
            logging.info(f"  --  {fname}: skipped for Specialty sheet")
            print(f"  --  {fname}: skipped for Specialty sheet")

        # ── Score ──────────────────────────────────────────────────────────────
        ok_score, _ = is_or_score_file(fpath)
        if ok_score:
            row = process_score_file(fpath, key_map, unknown_codes)
            if row is not None:
                score_rows.append(row)
        else:
            logging.info(f"  --  {fname}: skipped for Score sheet")

    spec_df = (
        pd.DataFrame(spec_rows, columns=SPEC_OUTPUT_COLS)
        if spec_rows
        else pd.DataFrame(columns=SPEC_OUTPUT_COLS)
    )
    score_df = (
        pd.DataFrame(score_rows, columns=SCORE_OUTPUT_COLS)
        if score_rows
        else pd.DataFrame(columns=SCORE_OUTPUT_COLS)
    )

    return spec_df, score_df, unknown_codes


# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT — WRITE BOTH SHEETS TO ONE WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

def _style_sheet(ws, df: pd.DataFrame, col_widths: list[int]) -> None:
    """Apply header + alternating-row styling to a worksheet."""
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 55

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    data_font = Font(name="Arial", size=9)
    alt_fill  = PatternFill("solid", start_color="EBF3FB")
    for row_num in range(2, len(df) + 2):
        for col_num in range(1, len(df.columns) + 1):
            cell           = ws.cell(row=row_num, column=col_num)
            cell.font      = data_font
            cell.alignment = Alignment(vertical="center")
            if row_num % 2 == 0:
                cell.fill = alt_fill

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_output(
    spec_df:  pd.DataFrame,
    score_df: pd.DataFrame,
    key_df:   pd.DataFrame | None,
    output_path: str,
) -> None:
    """Write both DataFrames to a single Excel workbook with two data sheets."""
    print(f"\nWriting output to: {output_path}")
    logging.info(
        f"Writing {len(spec_df)} specialty rows and {len(score_df)} score rows to {output_path}"
    )

    with pd.ExcelWriter(
        output_path, engine="openpyxl", datetime_format="DD-MMM-YYYY"
    ) as writer:
        # Sheet 1 — Specialty Level Data
        spec_df.to_excel(writer, sheet_name="Specialty Level Data", index=False)

        # Sheet 2 — Score
        score_df.to_excel(writer, sheet_name="Score", index=False)

        # Sheet 3 — Key (optional reference)
        if key_df is not None and len(key_df) > 0:
            key_df.to_excel(writer, sheet_name="Key", index=False)

        wb = writer.book

        # Style Sheet 1
        spec_widths = [22, 14, 32, 14, 28, 12, 10, 10, 10, 10, 10, 8, 8, 10, 10, 10, 14, 14, 10, 15, 10, 10, 10]
        _style_sheet(wb["Specialty Level Data"], spec_df, spec_widths)

        # Style Sheet 2
        score_widths = [22, 14, 30, 12, 8, 28, 8, 8, 10, 10, 10, 10, 14, 16, 18, 14, 18, 20, 18, 14, 18, 20, 18, 20]
        _style_sheet(wb["Score"], score_df, score_widths)


def write_unknown_hospitals(unknown_codes: set, output_path: str) -> None:
    if not unknown_codes:
        return
    unknown_path = os.path.splitext(output_path)[0] + "_UNKNOWN_HOSPITALS.txt"
    with open(unknown_path, "w") as f:
        f.write("Hospital codes not found in the Key table\n")
        f.write("=" * 45 + "\n")
        f.write("These hospitals will have a blank Directorate in the output.\n")
        f.write("Add them to the _EMBEDDED_KEY_TSV block in or_aggregator.py.\n\n")
        for code in sorted(unknown_codes):
            f.write(f"  {code}\n")
    print(f"\n  UNKNOWN HOSPITALS written to: {unknown_path}")
    logging.warning(f"Unknown hospitals file: {unknown_path}")


def print_summary(
    spec_df: pd.DataFrame,
    score_df: pd.DataFrame,
    unknown_codes: set,
    output_path: str,
) -> None:
    spec_filled  = (spec_df["Directorate"].notna() & (spec_df["Directorate"] != "")).sum()
    score_filled = (score_df["Directorate"].notna() & (score_df["Directorate"] != "")).sum() if len(score_df) else 0
    lines = [
        "=" * 60,
        f"  [Sheet 1 — Specialty Level Data]",
        f"    Rows written        : {len(spec_df):,}",
        f"    Hospitals processed : {spec_df['Hospital Code'].nunique() if len(spec_df) else 0}",
        f"    Directorates filled : {spec_filled:,} / {len(spec_df):,} rows",
        f"  [Sheet 2 — Score]",
        f"    Rows written        : {len(score_df):,}",
        f"    Hospitals processed : {score_df['Hospital Code'].nunique() if len(score_df) else 0}",
        f"    Directorates filled : {score_filled:,} / {len(score_df):,} rows",
    ]
    if unknown_codes:
        lines.append(f"  UNKNOWN hospitals   : {sorted(unknown_codes)}")
    lines += [f"  Saved to            : {output_path}", "=" * 60]
    summary = "\n".join(lines)
    print("\n" + summary)
    logging.info("\n" + summary)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description=(
            "OR Aggregator — processes OR data collector files and produces a single "
            "Excel workbook with two sheets: 'Specialty Level Data' and 'Score'."
        )
    )
    parser.add_argument("--input",  "-i", help="Folder containing raw OR data collector files")
    parser.add_argument(
        "--output", "-o", default="OR_Aggregated.xlsx",
        help="Output Excel file (default: OR_Aggregated.xlsx)"
    )
    parser.add_argument(
        "--key", "-k",
        help="(Optional) Excel file with a Key sheet to override the built-in mapping"
    )
    args = parser.parse_args()

    # ── Resolve input folder ──────────────────────────────────────────────────
    folder = args.input
    if not folder:
        folder = input("Enter the folder path containing OR data files:\n> ").strip().strip('"\'')
    if not os.path.isdir(folder):
        print(f"Error: folder not found: {folder}")
        sys.exit(1)

    output_path = args.output
    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"

    # ── Run log ───────────────────────────────────────────────────────────────
    log_path = os.path.splitext(output_path)[0] + "_run_log.txt"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(log_path, mode="w", encoding="utf-8"),
        ],
    )
    logging.info("OR Aggregator started")
    logging.info(f"Input folder : {folder}")
    logging.info(f"Output file  : {output_path}")

    # ── Key table ─────────────────────────────────────────────────────────────
    key_map, key_df = get_embedded_key()
    print(f"\nBuilt-in Key table loaded: {len(key_map)} hospitals across 21 directorates")
    logging.info(f"Embedded key: {len(key_map)} hospitals")

    ext_map, ext_df = load_key_from_file(key_path=args.key, search_folder=folder)
    if ext_map:
        key_map.update(ext_map)
        key_df = ext_df
        print(f"  External key merged — total {len(key_map)} hospital mappings")

    # ── Hospital name map ─────────────────────────────────────────────────────
    name_map = get_hospital_name_map()
    print(f"  Hospital name map loaded: {len(name_map)} entries")
    logging.info(f"Hospital name map: {len(name_map)} entries")

    # ── Process all files ─────────────────────────────────────────────────────
    spec_df, score_df, unknown_codes = aggregate_folder(folder, key_map, name_map=name_map)

    if spec_df.empty and score_df.empty:
        print("Nothing to write. Exiting.")
        sys.exit(0)

    # ── Write output ──────────────────────────────────────────────────────────
    write_output(spec_df, score_df, key_df, output_path)
    write_unknown_hospitals(unknown_codes, output_path)
    print_summary(spec_df, score_df, unknown_codes, output_path)

    logging.info("Run completed successfully")
    print(f"\n  Run log saved to: {log_path}")
    print("\nDone!")


if __name__ == "__main__":
    main()
